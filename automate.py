from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import openpyxl
import time


service = Service('chromedriver.exe')  # Update with your path to ChromeDriver
driver = webdriver.Chrome(service=service)
excel_path = 'input.xlsx'
workbook = openpyxl.load_workbook(excel_path)
sheet = workbook.active

driver.get("https://link.com")

time.sleep(10)

# Find the email input field and enter your email
username_input = driver.find_element(By.ID, "username")
username_input.send_keys("your_username")

# Find the password input field and enter your password
password_input = driver.find_element(By.ID, "password")
password_input.send_keys("your_password")

# Click the login button
login_button = driver.find_element(By.ID, "loginSubImg1")
login_button.click()

# Give it time to load after login
time.sleep(10)

# Navigate to "MANAGE PI"
driver.find_element(By.ID, 'switchPiTabLink').click()  # Replace with actual link text

'''
 # Select "Users" option
driver.find_element(By.LINK_TEXT, 'Users').click()  # Adjust as necessary
time.sleep(2) '''


driver.find_element(By.ID, 'displayAllUser').click() 
# Adjust as 

time.sleep(20)

#new feature start from here

# for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming first row is a header
#     username = row[0]  # Assuming usernames are in the first column

#     # Find the user in the table/list of users
#     # Adjust to locate users based on the structure of the page (e.g., table rows, divs, etc.)
#     users = driver.find_elements(By.XPATH, "//tr[contains(@class, 'list_separator list_user_separator')]")  # Update with actual class or element type
  

#     for user in users:
#         # Check if the username in the row matches the username in your Excel file
#         user_name_element = user.find_element(By.XPATH, ".//td/span[contains(@class, 'list2rowgreen')]")  # Update with actual username column class
#         if user_name_element.text == username:
            
            
#             print(f"Password updated for user: {username}")
        
#             break  # Exit the inner loop once the user is found and password is updated
        
#         else:
#     # This block runs if no break occurs, meaning the username was not found
#              print(f"Username not found: {username}")    
# time.sleep(200)



#new 2 fearture start from here
# Define the default and admin passwords
default_password = "Password"
admin_password = "password" 

updated_usernames = []

for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming first row is a header
    username = row[0]  # Assuming usernames are in the first column

    # Find the users in the table/list of users
    users = driver.find_elements(By.XPATH, "")  # Update with actual class or element type
    username_found = False
    # Iterate through the users to find the matching username
    for user in users:
        # Locate the username element inside the user row
        user_name_element = user.find_element(By.XPATH, "")  # Adjust with the actual structure
        
        # Check if the username in the table matches the username from the Excel file
        if user_name_element.text == username:
            time.sleep(5)
            #new 3 feater start from here
            # reset_password_link = user.find_element(By.XPATH, "//td/div[@class, 'editDeleteLink more-container')]")  # Update with actual link or button element
            # reset_password_link.click()
            # time.sleep(5)
            
            # Locate the div with the class 'editDeleteLink more-container' inside the td
            more_button = driver.find_element(By.XPATH, "")

# Click the div
            more_button.click()

# Optional: wait to see the result
            time.sleep(2)

            password_link = driver.find_element(By.XPATH, "")
            password_link.click()
            
            
            # Enter the new default password
            new_password_input = driver.find_element(By.ID, '')
            new_password_input.send_keys(default_password)
            time.sleep(5)
            # Confirm the default password
            # confirm_password_input = driver.find_element(By.ID, 'confirmPassword')
            # confirm_password_input.send_keys(default_password)
            
            update_password_input = driver.find_element(By.ID, 'verify_admin')
            update_password_input.click()
            time.sleep(5)
            # Enter the admin password for confirmation
            admin_password_input = driver.find_element(By.ID, 'admin_password')
            admin_password_input.send_keys(admin_password)
            time.sleep(5)
            # Submit the password change
            submit_button = driver.find_element(By.ID, 'update_credentials')  # Update with actual element ID
            submit_button.click()
            time.sleep(5)
            
            # Find the div element first (if necessary)
            div_element = driver.find_element(By.CLASS_NAME, '')

            # Now find the button within the div
            button = div_element.find_element(By.CLASS_NAME, '')

            # Click the button
            button.click()

            # Perform the password update for the matching user
            print(f"Password updated for user: {username}")
            updated_usernames.append(username)
            
            # Mark the username as found
            username_found = True

            #break  # Exit the loop once the user is found and the password is updated
    if not username_found:
        print(f"Username not found: {username}")

# After all usernames have been processed, print the final message
print("Updated all usernames.")

# Optionally, you can also print the list of updated usernames
print("The following usernames were updated:")
print(updated_usernames)
# Wait for some time (for testing or debugging purposes)
# time.sleep(200)




# Close the browser after performing actions
driver.quit() 